from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import desc
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .db import get_db, engine, Base
from .deps import get_current_user
from .models import ProductionReport, Product, ProductionReportAction

router = APIRouter(prefix="/production-reports", tags=["production-reports-export"])

def autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

@router.get("/export.xlsx")
def export_xlsx(
    from_date: date,
    to_date: date,
    status: str = "APPROVED",           # APPROVED / SUBMITTED / REJECTED / ALL
    group: str = "employee",            # employee / product / product_spec
    bucket: str = "day",                # day / week / month / year
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Base.metadata.create_all(bind=engine)

    # 拉取報表範圍內的 reports（details 用）
    q = (
        db.query(ProductionReport)
        .filter(ProductionReport.report_date >= from_date, ProductionReport.report_date <= to_date)
        .order_by(desc(ProductionReport.id))
    )
    if status != "ALL":
        q = q.filter(ProductionReport.status == status)
    reports = q.all()

    prod_map = {p.id: p for p in db.query(Product).all()}

    wb = Workbook()
    wb.remove(wb.active)

    # ===== Details =====
    ws_d = wb.create_sheet("Details")
    ws_d.append([
        "pr_no", "report_date", "status",
        "reported_by_user_id", "approved_by_user_id", "approved_at",
        "product_sku", "product_name", "product_spec", "spec_text",
        "qty", "unit", "item_note", "report_note",
    ])

    for pr in reports:
        for it in pr.items:
            p = prod_map.get(it.product_id)
            ws_d.append([
                pr.pr_no,
                str(pr.report_date),
                pr.status,
                pr.reported_by_user_id,
                pr.approved_by_user_id,
                str(pr.approved_at) if pr.approved_at else "",
                (p.sku if p else ""),
                (p.name if p else f"#{it.product_id}"),
                (getattr(p, "spec", "") if p else ""),
                (it.spec_text or ""),
                it.qty,
                it.unit,
                (it.note or ""),
                (pr.note or ""),
            ])
    autosize(ws_d)

    # ===== Rejected ===== (不論 status=ALL/APPROVED，都附一張退回表方便追)
    ws_r = wb.create_sheet("Rejected")
    ws_r.append(["pr_no", "report_date", "reported_by_user_id", "approved_by_user_id", "approved_at", "note"])
    rejected = (
        db.query(ProductionReport)
        .filter(ProductionReport.report_date >= from_date, ProductionReport.report_date <= to_date)
        .filter(ProductionReport.status == "REJECTED")
        .order_by(desc(ProductionReport.id))
        .all()
    )
    for pr in rejected:
        ws_r.append([pr.pr_no, str(pr.report_date), pr.reported_by_user_id, pr.approved_by_user_id, str(pr.approved_at or ""), pr.note or ""])
    autosize(ws_r)

    # ===== Summary =====（用 Details 在 Python 端彙總，避免 SQL 方言差異）
    # bucket key
    import datetime as _dt
    from collections import defaultdict

    def bucket_key(d: date) -> str:
        if bucket == "day":
            return d.strftime("%Y-%m-%d")
        if bucket == "week":
            # ISO week
            y, w, _ = d.isocalendar()
            return f"{y}-W{w:02d}"
        if bucket == "month":
            return d.strftime("%Y-%m")
        if bucket == "year":
            return d.strftime("%Y")
        return d.strftime("%Y-%m-%d")

    agg = defaultdict(int)  # (bucket, key, label) -> total_qty

    for pr in reports:
        # 報表只統計 APPROVED（除非 status=ALL 或 status!=APPROVED 你刻意要看）
        # 這裡尊重 status 參數：上面 reports 已依 status 篩過
        b = bucket_key(pr.report_date)

        for it in pr.items:
            p = prod_map.get(it.product_id)
            sku = p.sku if p and p.sku else ""
            pname = p.name if p else f"#{it.product_id}"
            pspec = getattr(p, "spec", "") if p else ""

            if group == "employee":
                key = f"employee:{pr.reported_by_user_id}"
                label = f"員工ID {pr.reported_by_user_id}"
            elif group == "product":
                key = f"product:{it.product_id}"
                label = f"{sku+' - ' if sku else ''}{pname}"
            else:  # product_spec
                spec = (it.spec_text or pspec or "-").strip() or "-"
                key = f"spec:{it.product_id}:{spec}"
                label = f"{sku+' - ' if sku else ''}{pname} | {spec}"

            agg[(b, key, label)] += int(it.qty or 0)

    ws_s = wb.create_sheet("Summary")
    ws_s.append(["bucket", "key", "label", "total_qty"])
    for (b, key, label), total in sorted(agg.items(), key=lambda x: (x[0][0], x[0][2])):
        ws_s.append([b, key, label, total])
    autosize(ws_s)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    # 記錄稽核：誰什麼時候匯出了什麼資料
    db.add(ProductionReportAction(
        report_id=0,  # 匯出不是針對單一 report，用 0 表示
        action="EXPORT",
        actor_user_id=user.id,
        actor_role=user.role.value if hasattr(user.role, 'value') else str(user.role),
        comment=f"from={from_date} to={to_date} status={status} group={group} bucket={bucket}",
    ))
    db.commit()

    filename = f"production_{from_date}_{to_date}_{status}_{group}_{bucket}.xlsx"
    return Response(
        content=out.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


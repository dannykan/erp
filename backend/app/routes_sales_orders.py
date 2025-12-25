from fastapi import APIRouter, Depends, HTTPException, Response, Query
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta, date, time
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .db import get_db, engine, Base
from .models import SalesOrder, SalesOrderItem, InventoryMove, MoveType, Product, Role, Customer
from .schemas import SOCreate, SOOut, SOView, SOCommonItemRow, SOPaged
from .deps import get_current_user, require_roles
from .utils import make_no
from .pdf_docs import build_so_pdf, build_so_picklist_pdf, build_so_shipping_pdf
from .constants import Site, Stage, RefType
from pydantic import BaseModel as PydanticBaseModel

router = APIRouter(prefix="/sales-orders", tags=["sales-orders"])

def so_to_view(db: Session, so: SalesOrder) -> SOView:
    from .models import Customer
    prod_map = {p.id: p for p in db.query(Product).all()}
    
    # 獲取客戶資訊（地址和電話）
    customer = db.query(Customer).filter(Customer.name == so.customer_name).first()
    customer_address = customer.address if customer else None
    customer_phone = customer.phone if customer else None
    
    items = []
    for it in so.items:
        p = prod_map.get(it.product_id)
        items.append({
            "id": it.id,
            "product_id": it.product_id,
            "product_sku": getattr(p, "sku", None) if p else None,
            "product_name": getattr(p, "name", f"#{it.product_id}") if p else f"#{it.product_id}",
            "product_spec": getattr(p, "spec", None) if p else None,
            "qty": float(it.qty),
            "unit": it.unit,
            "unit_price": float(getattr(it, "unit_price", 0) or 0),
            "price_unit": getattr(it, "price_unit", "件") or "件",
            "pieces_per_case": getattr(p, "pieces_per_case", None) if p else None,
            "mark": getattr(it, "mark", None) or "",
            "note": it.note or "",
        })
    return SOView(
        id=so.id,
        so_no=so.so_no,
        customer_name=so.customer_name,
        customer_address=customer_address,
        customer_phone=customer_phone,
        doc_date=so.doc_date,
        note=so.note,
        status=so.status,
        created_at=so.created_at,
        picked_at=so.picked_at,
        picked_by_id=so.picked_by_id,
        shipped_at=so.shipped_at,
        shipped_by_id=so.shipped_by_id,
        ship_note=so.ship_note,
        logistics_no=so.logistics_no,
        is_paid=getattr(so, "is_paid", False),
        paid_at=getattr(so, "paid_at", None),
        paid_by_id=getattr(so, "paid_by_id", None),
        items=items,
    )

@router.get("", response_model=SOPaged)
def list_sos(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),

    # filters
    date_from: Optional[date] = Query(None),  # 保留兼容性，但优先使用 shipped_at_from
    date_to: Optional[date] = Query(None),    # 保留兼容性，但优先使用 shipped_at_to
    shipped_at_from: Optional[date] = Query(None, description="出貨時間起始日期"),
    shipped_at_to: Optional[date] = Query(None, description="出貨時間結束日期"),
    customer_name_like: Optional[str] = Query(None),
    status: Optional[str] = Query(None),  # DRAFT/PICKED/SHIPPED
    product_id: Optional[int] = Query(None),
    is_paid: Optional[bool] = Query(None, description="付款狀態篩選：True=已付款, False=未付款"),

    # paging
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    Base.metadata.create_all(bind=engine)

    q = db.query(SalesOrder)

    # 優先使用出貨時間篩選，如果沒有則使用 doc_date（向後兼容）
    if shipped_at_from:
        q = q.filter(SalesOrder.shipped_at >= datetime.combine(shipped_at_from, time.min))
    elif date_from:
        q = q.filter(SalesOrder.doc_date >= date_from)
    
    if shipped_at_to:
        q = q.filter(SalesOrder.shipped_at <= datetime.combine(shipped_at_to, time.max))
    elif date_to:
        q = q.filter(SalesOrder.doc_date <= date_to)

    if customer_name_like:
        kw = customer_name_like.strip()
        if kw:
            q = q.filter(SalesOrder.customer_name.ilike(f"%{kw}%"))

    if status:
        q = q.filter(SalesOrder.status == status)

    # 付款狀態篩選：只對已出貨的單據有效
    if is_paid is not None:
        q = q.filter(SalesOrder.status == 'SHIPPED')
        q = q.filter(SalesOrder.is_paid == is_paid)

    if product_id:
        q = q.join(SalesOrderItem, SalesOrderItem.sales_order_id == SalesOrder.id)\
             .filter(SalesOrderItem.product_id == product_id)\
             .distinct()

    total = q.count()

    # 排序：優先按出貨時間倒序，沒有出貨時間的按建立時間倒序
    rows = (
        q.order_by(
            desc(SalesOrder.shipped_at),
            desc(SalesOrder.created_at),
            desc(SalesOrder.id)
        )
         .offset((page - 1) * page_size)
         .limit(page_size)
         .all()
    )

    return {
        "rows": [so_to_view(db, so) for so in rows],
        "total": total,
    }

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

@router.get("/export.xlsx")
def export_sos_xlsx(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),

    date_from: Optional[date] = Query(None),  # 保留兼容性，但优先使用 shipped_at_from
    date_to: Optional[date] = Query(None),    # 保留兼容性，但优先使用 shipped_at_to
    shipped_at_from: Optional[date] = Query(None, description="出貨時間起始日期"),
    shipped_at_to: Optional[date] = Query(None, description="出貨時間結束日期"),
    customer_name_like: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    product_id: Optional[int] = Query(None),
    is_paid: Optional[bool] = Query(None),
):
    # 這裡重用同一套 query 邏輯（跟 list_sos 一致）
    q = db.query(SalesOrder)

    # 優先使用出貨時間篩選，如果沒有則使用 doc_date（向後兼容）
    if shipped_at_from:
        q = q.filter(SalesOrder.shipped_at >= datetime.combine(shipped_at_from, time.min))
    elif date_from:
        q = q.filter(SalesOrder.doc_date >= date_from)
    
    if shipped_at_to:
        q = q.filter(SalesOrder.shipped_at <= datetime.combine(shipped_at_to, time.max))
    elif date_to:
        q = q.filter(SalesOrder.doc_date <= date_to)

    if customer_name_like:
        kw = customer_name_like.strip()
        if kw:
            q = q.filter(SalesOrder.customer_name.ilike(f"%{kw}%"))

    if status:
        q = q.filter(SalesOrder.status == status)

    # 付款狀態篩選：只對已出貨的單據有效
    if is_paid is not None:
        q = q.filter(SalesOrder.status == 'SHIPPED')
        q = q.filter(SalesOrder.is_paid == is_paid)

    if product_id:
        q = q.join(SalesOrderItem, SalesOrderItem.sales_order_id == SalesOrder.id)\
             .filter(SalesOrderItem.product_id == product_id)\
             .distinct()

    # 排序：優先按出貨時間倒序，沒有出貨時間的按建立時間倒序
    sos = q.order_by(
        desc(SalesOrder.shipped_at),
        desc(SalesOrder.created_at),
        desc(SalesOrder.id)
    ).all()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws2 = wb.create_sheet("Items")

    ws1.append(["so_no", "doc_date", "customer_name", "status", "items_count", "total_amount", "note"])
    ws2.append(["so_no", "doc_date", "customer_name", "status", "product_id", "product_name", "qty", "unit", "unit_price", "price_unit", "line_total", "note"])

    for so in sos:
        view = so_to_view(db, so)  # 你既有轉換（含 items, product_name 等）
        items = getattr(view, "items", []) or []
        total_amount = 0.0
        
        # Format doc_date once per order
        doc_date_str = ""
        doc_date_val = getattr(view, "doc_date", None)
        if doc_date_val:
            doc_date_str = str(doc_date_val)
        
        for it in items:
            qty = float(getattr(it, "qty", 0) or 0)
            unit_price = float(getattr(it, "unit_price", 0) or 0)
            line_total = qty * unit_price
            total_amount += line_total

            ws2.append([
                getattr(view, "so_no", ""),
                doc_date_str,
                getattr(view, "customer_name", ""),
                getattr(view, "status", ""),
                getattr(it, "product_id", None),
                getattr(it, "product_name", ""),  # 你前面已做過「明細顯示商品名稱」
                qty,
                getattr(it, "unit", ""),
                unit_price,
                getattr(it, "price_unit", "") or getattr(it, "unit", "") or "件",
                line_total,
                getattr(it, "note", ""),
            ])

        ws1.append([
            getattr(view, "so_no", ""),
            doc_date_str,
            getattr(view, "customer_name", ""),
            getattr(view, "status", ""),
            len(items),
            total_amount,
            getattr(view, "note", ""),
        ])

    _autosize(ws1)
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "sales_orders_export.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.post("", response_model=SOView)
def create_so(
    payload: SOCreate,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    if not payload.items:
        raise HTTPException(400, "items required")

    # 檢查商品存在並驗證資料
    for it in payload.items:
        if not it.product_id:
            raise HTTPException(400, "product_id is required for all items")
        p = db.query(Product).filter(Product.id == it.product_id).first()
        if not p:
            raise HTTPException(400, f"product_id {it.product_id} not found")
        if it.qty is None or it.qty < 0:
            raise HTTPException(400, f"qty must be >= 0 for product_id {it.product_id}")

    prefix = make_no("SO")
    seq = db.query(SalesOrder).filter(SalesOrder.so_no.like(f"{prefix}-%")).count() + 1
    so_no = f"{prefix}-{seq:04d}"

    # Normalize customer_name: strip whitespace
    customer_name = payload.customer_name.strip()

    so = SalesOrder(
        so_no=so_no,
        customer_name=customer_name,
        doc_date=payload.doc_date,
        note=payload.note,
        status="DRAFT",
    )

    # 建立 items + 寫入庫存 OUT 流水（建立時就扣庫，允許庫存為負數）
    for it in payload.items:
        # qty 在 model 中是 int，但 schema 允許 float，所以轉換為 int
        qty_val = int(it.qty) if isinstance(it.qty, (int, float)) and it.qty > 0 else 0
        if qty_val <= 0:
            continue  # 跳過數量為 0 或負數的項目（這些項目不應該建立）
        
        # Normalize unit and price_unit
        unit_val = (it.unit or "包").strip() if it.unit else "包"
        price_unit_val = (it.price_unit or it.unit or "件").strip() if (it.price_unit or it.unit) else "件"
        
        so.items.append(SalesOrderItem(
            product_id=it.product_id,
            qty=qty_val,
            unit=unit_val,
            unit_price=float(it.unit_price or 0),
            price_unit=price_unit_val,
            note=it.note,
            mark=it.mark,
        ))
        
        # 建立銷貨單時就扣除庫存（允許庫存變為負數）
        db.add(InventoryMove(
            product_id=it.product_id,
            move_type=MoveType.OUT,
            qty_change=-float(qty_val),
            site=Site.WAREHOUSE,
            stage="DRAFT",  # 建立銷貨單時的階段
            ref_type=RefType.SO,
            ref_no=so_no,
            note=f"SO {so_no} 建立銷貨單",
        ))

    # 確保至少有一個有效的項目
    if not so.items:
        raise HTTPException(400, "至少需要一個有效的項目（數量 > 0）")
    
    db.add(so)
    try:
        db.commit()
        db.refresh(so)
        return so_to_view(db, so)
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = str(e)
        traceback.print_exc()
        raise HTTPException(500, f"建立銷貨單失敗：{error_detail}")

@router.put("/{so_id}", response_model=SOView)
def update_so(
    so_id: int,
    payload: SOCreate,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "Sales order not found")
    
    # 只有 DRAFT 狀態的銷貨單可以編輯
    if so.status != "DRAFT":
        raise HTTPException(400, f"只有待出貨狀態的銷貨單可以編輯，目前狀態：{so.status}")
    
    if not payload.items:
        raise HTTPException(400, "items required")

    # 檢查商品存在並驗證資料
    for it in payload.items:
        if not it.product_id:
            raise HTTPException(400, "product_id is required for all items")
        p = db.query(Product).filter(Product.id == it.product_id).first()
        if not p:
            raise HTTPException(400, f"product_id {it.product_id} not found")
        if it.qty is None or it.qty < 0:
            raise HTTPException(400, f"qty must be >= 0 for product_id {it.product_id}")

    # 刪除舊的庫存異動記錄（建立時扣除的庫存）
    old_moves = db.query(InventoryMove).filter(
        InventoryMove.ref_type == RefType.SO,
        InventoryMove.ref_no == so.so_no,
        InventoryMove.stage == "DRAFT",
    ).all()
    for mv in old_moves:
        db.delete(mv)
    
    # 刪除舊的 items
    for item in so.items:
        db.delete(item)
    so.items.clear()
    
    # Normalize customer_name: strip whitespace
    customer_name = payload.customer_name.strip()
    
    # 更新基本資訊
    so.customer_name = customer_name
    so.doc_date = payload.doc_date
    so.note = payload.note

    # 建立新的 items + 寫入庫存 OUT 流水（建立時就扣庫，允許庫存為負數）
    for it in payload.items:
        # qty 在 model 中是 int，但 schema 允許 float，所以轉換為 int
        qty_val = int(it.qty) if isinstance(it.qty, (int, float)) and it.qty > 0 else 0
        if qty_val <= 0:
            continue  # 跳過數量為 0 或負數的項目（這些項目不應該建立）
        
        # Normalize unit and price_unit
        unit_val = (it.unit or "包").strip() if it.unit else "包"
        price_unit_val = (it.price_unit or it.unit or "件").strip() if (it.price_unit or it.unit) else "件"
        
        so.items.append(SalesOrderItem(
            product_id=it.product_id,
            qty=qty_val,
            unit=unit_val,
            unit_price=float(it.unit_price or 0),
            price_unit=price_unit_val,
            note=it.note,
            mark=it.mark,
        ))
        
        # 建立銷貨單時就扣除庫存（允許庫存變為負數）
        db.add(InventoryMove(
            product_id=it.product_id,
            move_type=MoveType.OUT,
            qty_change=-float(qty_val),
            site=Site.WAREHOUSE,
            stage="DRAFT",  # 建立銷貨單時的階段
            ref_type=RefType.SO,
            ref_no=so.so_no,
            note=f"SO {so.so_no} 建立銷貨單",
        ))

    # 確保至少有一個有效的項目
    if not so.items:
        raise HTTPException(400, "至少需要一個有效的項目（數量 > 0）")
    
    try:
        db.commit()
        db.refresh(so)
        return so_to_view(db, so)
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = str(e)
        traceback.print_exc()
        raise HTTPException(500, f"更新銷貨單失敗：{error_detail}")

@router.get("/common-items", response_model=list[SOCommonItemRow])
def so_common_items(
    customer_name: str = Query(..., min_length=1),
    limit: int = Query(50, ge=1, le=200),
    days: int = Query(180, ge=1, le=3650),
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    cname = customer_name.strip()
    if not cname:
        raise HTTPException(status_code=400, detail="customer_name cannot be empty")
    since = datetime.utcnow() - timedelta(days=days)

    # freq per product
    freq_rows = (
        db.query(
            SalesOrderItem.product_id.label("product_id"),
            func.count(SalesOrderItem.id).label("freq"),
            func.max(SalesOrder.created_at).label("last_ts"),
        )
        .join(SalesOrder, SalesOrder.id == SalesOrderItem.sales_order_id)
        .filter(SalesOrder.customer_name == cname)
        .filter(SalesOrder.created_at >= since)
        .group_by(SalesOrderItem.product_id)
        .order_by(func.count(SalesOrderItem.id).desc(), func.max(SalesOrder.created_at).desc())
        .limit(limit)
        .all()
    )

    if not freq_rows:
        return []

    product_ids = [r.product_id for r in freq_rows]

    # last row per product: pick by last_ts + join (use correlated subquery to be safe)
    # simplest: for each product_id, query latest item (N+1) — OK for limit <= 50
    prod_map = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()}

    out = []
    for r in freq_rows:
        last_item = (
            db.query(SalesOrderItem, SalesOrder)
            .join(SalesOrder, SalesOrder.id == SalesOrderItem.sales_order_id)
            .filter(SalesOrder.customer_name == cname)
            .filter(SalesOrderItem.product_id == r.product_id)
            .order_by(SalesOrder.created_at.desc(), SalesOrderItem.id.desc())
            .first()
        )
        if not last_item:
            continue
        it, so = last_item
        p = prod_map.get(r.product_id)
        if not p:
            continue
        last_price_unit = getattr(it, "price_unit", None) or getattr(it, "unit", None) or "件"
        out.append(SOCommonItemRow(
            product_id=r.product_id,
            sku=getattr(p, "sku", None),
            name=p.name,
            last_unit_price=float(getattr(it, "unit_price", 0) or 0),
            last_price_unit=last_price_unit,
            last_qty=float(it.qty),
            last_order_date=so.doc_date,
            freq=int(r.freq),
        ))
    return out

@router.get("/last", response_model=SOView)
def last_so(
    customer_name: str = Query(...),
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    cname = customer_name.strip()
    so = (
        db.query(SalesOrder)
        .filter(SalesOrder.customer_name == cname)
        .order_by(desc(SalesOrder.created_at))
        .first()
    )
    if not so:
        raise HTTPException(404, "找不到該客戶的歷史訂單")
    return so_to_view(db, so)

@router.get("/{so_id}", response_model=SOView)
def get_so(
    so_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")
    return so_to_view(db, so)

@router.get("/{so_id}/print")
def print_so(
    so_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")

    # 取得客戶資訊
    customer = db.query(Customer).filter(Customer.name == so.customer_name).first()
    customer_address = customer.address if customer else None
    customer_phone = customer.phone if customer else None

    prod_map = {p.id: p for p in db.query(Product).all()}
    items = []
    for it in so.items:
        p = prod_map.get(it.product_id)
        items.append({
            "product_sku": p.sku if p else None,
            "product_name": p.name if p else f"#{it.product_id}",
            "product_spec": p.spec if p else None,
            "qty": it.qty,
            "case_qty": it.qty,  # 使用 qty 作為 case_qty
            "unit": it.unit,
            "unit_price": float(getattr(it, "unit_price", 0) or 0),
            "price_unit": getattr(it, "price_unit", "件") or "件",
            "pieces_per_case": p.pieces_per_case if p else None,
            "mark": "",  # MARK 欄位，目前為空
            "note": it.note,
        })

    pdf = build_so_pdf(so.so_no, so.customer_name, str(so.doc_date), items, so.note, customer_address, customer_phone)
    return Response(content=pdf, media_type="application/pdf")

@router.get("/{so_id}/picklist.pdf")
def print_picklist(
    so_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")
    # 狀態限制：只允許 PICKED/SHIPPED 下載（避免 DRAFT 亂印）
    if so.status not in ("PICKED", "SHIPPED"):
        raise HTTPException(400, "Picklist can only be printed for PICKED or SHIPPED orders")

    prod_map = {p.id: p for p in db.query(Product).all()}
    items = []
    for it in so.items:
        p = prod_map.get(it.product_id)
        items.append({
            "product_name": (f"{p.sku} - {p.name}" if p and p.sku else (p.name if p else f"#{it.product_id}")),
            "qty": it.qty,
            "unit": it.unit,
            "note": it.note,
        })

    pdf = build_so_picklist_pdf(so.so_no, so.customer_name, str(so.doc_date), items, so.note)
    return Response(content=pdf, media_type="application/pdf")

@router.get("/{so_id}/shipping.pdf")
def print_shipping(
    so_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")
    # 狀態限制：只允許 SHIPPED 下載（避免未出貨先印出貨單）
    if so.status != "SHIPPED":
        raise HTTPException(400, "Shipping document can only be printed for SHIPPED orders")

    prod_map = {p.id: p for p in db.query(Product).all()}
    items = []
    for it in so.items:
        p = prod_map.get(it.product_id)
        items.append({
            "product_name": (f"{p.sku} - {p.name}" if p and p.sku else (p.name if p else f"#{it.product_id}")),
            "qty": it.qty,
            "unit": it.unit,
            "unit_price": float(getattr(it, "unit_price", 0) or 0),
            "price_unit": getattr(it, "price_unit", "件") or "件",
            "note": it.note,
        })

    shipped_at_str = str(so.shipped_at) if so.shipped_at else None
    pdf = build_so_shipping_pdf(
        so.so_no, 
        so.customer_name, 
        str(so.doc_date), 
        items, 
        so.note,
        shipped_at_str,
        so.logistics_no,
        so.ship_note
    )
    return Response(content=pdf, media_type="application/pdf")

@router.post("/{so_id}/pick", response_model=SOView)
def pick_so(
    so_id: int,
    db: Session = Depends(get_db),
    # TODO: 未来如果有 warehouse 角色，应该允许 warehouse 角色执行 pick/ship
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")
    if so.status != "DRAFT":
        raise HTTPException(400, "Only DRAFT can be picked")

    so.status = "PICKED"
    so.picked_at = datetime.utcnow()
    so.picked_by_id = user.id
    db.commit()
    db.refresh(so)
    return so_to_view(db, so)

class SOShipIn(PydanticBaseModel):
    ship_note: Optional[str] = None
    logistics_no: Optional[str] = None

@router.post("/{so_id}/ship", response_model=SOView)
def ship_so(
    so_id: int,
    payload: SOShipIn = SOShipIn(),
    db: Session = Depends(get_db),
    # TODO: 未来如果有 warehouse 角色，应该允许 warehouse 角色执行 pick/ship
    # TODO: 建议添加 audit log（action=SHIP/PICK, actor_user_id, comment: so_no/logistics_no）
    # 用于追踪「谁出错」，类似 KPI/export 的 audit log
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    Base.metadata.create_all(bind=engine)
    
    # 使用 SELECT FOR UPDATE 鎖住 SO，防止并发出货
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).with_for_update().first()
    if not so:
        raise HTTPException(404, "SO not found")
    
    # idempotency：若已寫過出貨標記，直接回傳（放在 status 檢查前，避免重整頁面誤按時回 400）
    existed = db.query(InventoryMove).filter(
        InventoryMove.ref_type == RefType.SO,
        InventoryMove.ref_no == str(so_id),
        InventoryMove.stage == Stage.SHIP,
    ).first()
    if existed:
        db.refresh(so)
        return so_to_view(db, so)
    
    if so.status != "PICKED":
        raise HTTPException(400, "Only PICKED can be shipped")

    # 庫存已在建立銷貨單時扣除，出貨時只更新狀態，不再扣庫存
    # 檢查是否已經有庫存異動記錄（建立時已扣庫）
    existing_moves = db.query(InventoryMove).filter(
        InventoryMove.ref_type == RefType.SO,
        InventoryMove.ref_no == so.so_no,
        InventoryMove.stage == "DRAFT",
    ).all()
    
    # 如果沒有庫存異動記錄（可能是舊的銷貨單），則在出貨時補扣庫存
    if not existing_moves:
        # 庫存檢查（使用浮點數 epsilon 避免精度問題）- 只提醒不阻止
        EPSILON = 1e-9
        low_stock_items = []
        for it in so.items:
            pid = it.product_id
            need = float(it.qty or 0)
            stock = db.query(
                func.coalesce(func.sum(InventoryMove.qty_change), 0)
            ).filter(
                InventoryMove.product_id == pid,
                InventoryMove.site == Site.WAREHOUSE
            ).scalar()
            have = float(stock or 0)
            if have + EPSILON < need:
                p = db.query(Product).filter(Product.id == pid).first()
                product_name = p.name if p else f"商品 #{pid}"
                low_stock_items.append(f"{product_name}（庫存：{have}，需求：{need}）")
        
        # 如果有庫存不足的商品，記錄警告但不阻止出貨
        if low_stock_items:
            import logging
            logging.warning(f"SO {so.so_no} shipped with low stock: {', '.join(low_stock_items)}")

        # 寫入庫存 move（WAREHOUSE 出庫）- 舊的銷貨單才需要
        for it in so.items:
            mv = InventoryMove(
                product_id=it.product_id,
                move_type=MoveType.OUT,
                qty_change=-float(it.qty or 0),
                site=Site.WAREHOUSE,
                stage=Stage.SHIP,
                ref_type=RefType.SO,
                ref_no=str(so_id),
                note=f"SO {so.so_no} SHIP",
                created_at=datetime.utcnow(),
            )
            db.add(mv)
    else:
        # 更新現有的庫存異動記錄的 stage 為 SHIP
        for mv in existing_moves:
            mv.stage = Stage.SHIP
            mv.note = f"SO {so.so_no} SHIP"

    so.status = "SHIPPED"
    so.shipped_at = datetime.utcnow()
    so.shipped_by_id = user.id
    so.ship_note = (payload.ship_note or "").strip() or None
    so.logistics_no = (payload.logistics_no or "").strip() or None

    # 所有操作在同一個 transaction 中 commit（原子性保證）
    # 如果遇到 unique constraint violation（并发出货），视为已出货
    try:
        db.commit()
        db.refresh(so)
        return so_to_view(db, so)
    except IntegrityError as e:
        # 只放行 unique index 的 violation，避免把其他 DB 错误也吞掉
        err_upper = str(e).upper()
        err_lower = str(e).lower()
        is_unique_violation = (
            "UNIQUE CONSTRAINT FAILED" in err_upper
            or "UNIQUE CONSTRAINT" in err_upper
            or "UNIQUE_INDEX" in err_upper
            or "uq_inventory_moves_ref_stage_site_product" in err_lower
        )
        
        if not is_unique_violation:
            # 不是 unique constraint 错误，重新抛出
            db.rollback()
            raise HTTPException(500, f"Database error: {str(e)}")
        
        # 是 unique constraint violation（并发出货）
        # 1. 先 rollback（确保 transaction 状态干净）
        db.rollback()
        # 重新查询 SO 状态（可能已被其他请求更新）
        so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
        if not so:
            raise HTTPException(404, "SO not found")
        
        # 检查是否已有出货 move（即使 status 可能还是 PICKED）
        existed_move = db.query(InventoryMove).filter(
            InventoryMove.ref_type == RefType.SO,
            InventoryMove.ref_no == str(so_id),
            InventoryMove.stage == Stage.SHIP,
        ).first()
        
        if existed_move and so.status == "PICKED":
            # 另一个交易已写入 moves 但还没更新 status，自动修复
            so.status = "SHIPPED"
            if not so.shipped_at:
                so.shipped_at = existed_move.created_at or datetime.utcnow()
            db.commit()
            db.refresh(so)
        
        db.refresh(so)
        return so_to_view(db, so)

@router.post("/{so_id}/confirm-payment", response_model=SOView)
def confirm_payment(
    so_id: int,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.office, Role.supervisor)),
):
    """确认收款 - 只有已出貨的銷貨單可以確認收款"""
    Base.metadata.create_all(bind=engine)
    so = db.query(SalesOrder).filter(SalesOrder.id == so_id).first()
    if not so:
        raise HTTPException(404, "SO not found")
    if so.status != "SHIPPED":
        raise HTTPException(400, "Only SHIPPED orders can confirm payment")
    if so.is_paid:
        raise HTTPException(400, "Order is already paid")

    so.is_paid = True
    so.paid_at = datetime.utcnow()
    so.paid_by_id = user.id

    db.commit()
    db.refresh(so)
    return so_to_view(db, so)


@router.get("/merged-unpaid")
def get_merged_unpaid_sos(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    customer_name: str = Query(..., description="客戶名稱"),
    shipped_at_from: Optional[date] = Query(None, description="出貨時間起始日期"),
    shipped_at_to: Optional[date] = Query(None, description="出貨時間結束日期"),
):
    """獲取某個客戶未付款的銷貨單，合併品項"""
    from collections import defaultdict
    
    # 查詢該客戶未付款的已出貨銷貨單
    q = db.query(SalesOrder).filter(
        SalesOrder.customer_name == customer_name,
        SalesOrder.status == 'SHIPPED',
        SalesOrder.is_paid == False
    )
    
    if shipped_at_from:
        q = q.filter(SalesOrder.shipped_at >= datetime.combine(shipped_at_from, time.min))
    if shipped_at_to:
        q = q.filter(SalesOrder.shipped_at <= datetime.combine(shipped_at_to, time.max))
    
    sos = q.order_by(SalesOrder.shipped_at).all()
    
    if not sos:
        return {
            "customer_name": customer_name,
            "customer_address": None,
            "customer_phone": None,
            "date_from": shipped_at_from.isoformat() if shipped_at_from else None,
            "date_to": shipped_at_to.isoformat() if shipped_at_to else None,
            "source_so_ids": [],
            "source_so_nos": [],
            "items": [],
            "total_amount": 0.0,
            "total_qty": 0.0,
        }
    
    # 獲取客戶資訊
    customer = db.query(Customer).filter(Customer.name == customer_name).first()
    customer_address = customer.address if customer else None
    customer_phone = customer.phone if customer else None
    
    # 合併品項：按 product_id + mark 分組
    merged_items: dict[tuple, dict] = defaultdict(lambda: {
        "product_id": 0,
        "product_sku": None,
        "product_name": "",
        "product_spec": None,
        "total_qty": 0.0,
        "unit": "",
        "unit_price": 0.0,
        "price_unit": "",
        "total_amount": 0.0,
        "source_so_nos": set(),
        "mark": None,
        "note": None,
    })
    
    prod_map = {p.id: p for p in db.query(Product).all()}
    source_so_ids = []
    source_so_nos = []
    
    for so in sos:
        source_so_ids.append(so.id)
        source_so_nos.append(so.so_no)
        
        for item in so.items:
            p = prod_map.get(item.product_id)
            # 使用 (product_id, mark) 作為合併鍵
            key = (item.product_id, item.mark or "")
            
            merged = merged_items[key]
            if merged["product_id"] == 0:
                merged["product_id"] = item.product_id
                merged["product_sku"] = getattr(p, "sku", None) if p else None
                merged["product_name"] = getattr(p, "name", f"#{item.product_id}") if p else f"#{item.product_id}"
                merged["product_spec"] = getattr(p, "spec", None) if p else None
                merged["unit"] = item.unit
                merged["unit_price"] = float(getattr(item, "unit_price", 0) or 0)
                merged["price_unit"] = getattr(item, "price_unit", "件") or "件"
                merged["mark"] = item.mark
                merged["note"] = item.note
            
            qty = float(item.qty)
            merged["total_qty"] += qty
            merged["total_amount"] += qty * merged["unit_price"]
            merged["source_so_nos"].add(so.so_no)
    
    # 轉換為列表格式
    items = []
    for key, merged in merged_items.items():
        items.append({
            "product_id": merged["product_id"],
            "product_sku": merged["product_sku"],
            "product_name": merged["product_name"],
            "product_spec": merged["product_spec"],
            "total_qty": merged["total_qty"],
            "unit": merged["unit"],
            "unit_price": merged["unit_price"],
            "price_unit": merged["price_unit"],
            "total_amount": merged["total_amount"],
            "source_so_nos": sorted(list(merged["source_so_nos"])),
            "mark": merged["mark"],
            "note": merged["note"],
        })
    
    total_amount = sum(item["total_amount"] for item in items)
    total_qty = sum(item["total_qty"] for item in items)
    
    return {
        "customer_name": customer_name,
        "customer_address": customer_address,
        "customer_phone": customer_phone,
        "date_from": shipped_at_from.isoformat() if shipped_at_from else None,
        "date_to": shipped_at_to.isoformat() if shipped_at_to else None,
        "source_so_ids": source_so_ids,
        "source_so_nos": sorted(source_so_nos),
        "items": items,
        "total_amount": total_amount,
        "total_qty": total_qty,
    }

@router.get("/merged-unpaid/print.pdf")
def print_merged_unpaid_sos(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    customer_name: str = Query(..., description="客戶名稱"),
    shipped_at_from: Optional[date] = Query(None, description="出貨時間起始日期"),
    shipped_at_to: Optional[date] = Query(None, description="出貨時間結束日期"),
):
    """列印合併的未付款銷貨單（陣列式格式）"""
    from .pdf_docs import build_merged_unpaid_so_pdf
    from collections import defaultdict
    
    q = db.query(SalesOrder).filter(
        SalesOrder.customer_name == customer_name,
        SalesOrder.status == 'SHIPPED',
        SalesOrder.is_paid == False
    )
    
    if shipped_at_from:
        q = q.filter(SalesOrder.shipped_at >= datetime.combine(shipped_at_from, time.min))
    if shipped_at_to:
        q = q.filter(SalesOrder.shipped_at <= datetime.combine(shipped_at_to, time.max))
    
    sos = q.order_by(SalesOrder.shipped_at).all()
    
    if not sos:
        raise HTTPException(404, "No unpaid orders found")
    
    customer = db.query(Customer).filter(Customer.name == customer_name).first()
    customer_address = customer.address if customer else None
    customer_phone = customer.phone if customer else None
    
    merged_items: dict[tuple, dict] = defaultdict(lambda: {
        "product_id": 0,
        "product_sku": None,
        "product_name": "",
        "product_spec": None,
        "total_qty": 0.0,
        "unit": "",
        "unit_price": 0.0,
        "price_unit": "",
        "total_amount": 0.0,
        "source_so_nos": set(),
        "mark": None,
        "note": None,
    })
    
    prod_map = {p.id: p for p in db.query(Product).all()}
    source_so_nos = []
    
    for so in sos:
        source_so_nos.append(so.so_no)
        for item in so.items:
            p = prod_map.get(item.product_id)
            key = (item.product_id, item.mark or "")
            merged = merged_items[key]
            if merged["product_id"] == 0:
                merged["product_id"] = item.product_id
                merged["product_sku"] = getattr(p, "sku", None) if p else None
                merged["product_name"] = getattr(p, "name", f"#{item.product_id}") if p else f"#{item.product_id}"
                merged["product_spec"] = getattr(p, "spec", None) if p else None
                merged["unit"] = item.unit
                merged["unit_price"] = float(getattr(item, "unit_price", 0) or 0)
                merged["price_unit"] = getattr(item, "price_unit", "件") or "件"
                merged["mark"] = item.mark
                merged["note"] = item.note
            qty = float(item.qty)
            merged["total_qty"] += qty
            merged["total_amount"] += qty * merged["unit_price"]
            merged["source_so_nos"].add(so.so_no)
    
    items_for_pdf = []
    total_amount = 0.0
    total_qty = 0.0
    for key, merged in merged_items.items():
        items_for_pdf.append({
            "product_sku": merged["product_sku"],
            "product_name": merged["product_name"],
            "product_spec": merged["product_spec"],
            "qty": merged["total_qty"],
            "unit": merged["unit"],
            "unit_price": merged["unit_price"],
            "price_unit": merged["price_unit"],
            "case_qty": merged["total_qty"],
            "pieces_per_case": getattr(prod_map.get(merged["product_id"]), "pieces_per_case", None) if merged["product_id"] else None,
            "mark": merged["mark"],
            "note": merged["note"],
        })
        total_amount += merged["total_amount"]
        total_qty += merged["total_qty"]
    
    pdf_bytes = build_merged_unpaid_so_pdf(
        customer_name=customer_name,
        customer_address=customer_address,
        customer_phone=customer_phone,
        date_from=shipped_at_from.isoformat() if shipped_at_from else None,
        date_to=shipped_at_to.isoformat() if shipped_at_to else None,
        source_so_nos=sorted(source_so_nos),
        items=items_for_pdf,
        total_amount=total_amount,
        total_qty=total_qty,
    )
    
    filename = f"merged_unpaid_{customer_name}_{shipped_at_from.isoformat() if shipped_at_from else ''}_{shipped_at_to.isoformat() if shipped_at_to else ''}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

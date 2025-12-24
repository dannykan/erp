from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .db import get_db, engine, Base
from .deps import require_roles
from .models import Role, ProductionReport, ProductionReportItem, Product, ProductionReportAction, User
from .schemas import ProductionKPIOut, KPIBlock, KPIRankRow

router = APIRouter(prefix="", tags=["production-kpi"])

def normalize_reason(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "（未填原因）"
    # 取第一句/前 50 字，避免太長
    s = re.split(r"[\n\r。；;]", s)[0].strip()
    return s[:50] if len(s) > 50 else s

def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

BOLD = Font(bold=True)

def _freeze_and_header(ws: Worksheet, header_row: int = 1):
    ws.freeze_panes = f"A{header_row+1}"
    for cell in ws[header_row]:
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _fmt_percent_cell(ws: Worksheet, row: int, col: int):
    ws.cell(row=row, column=col).number_format = '0.00"%"'

def _compute_kpi(db: Session, from_date: date, to_date: date, top_n: int) -> ProductionKPIOut:
    """核心 KPI 計算邏輯，可被路由和匯出函數重用"""
    Base.metadata.create_all(bind=engine)

    # --- totals ---
    total_all = db.query(func.count(ProductionReport.id)).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
    ).scalar() or 0

    total_rejected = db.query(func.count(ProductionReport.id)).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "REJECTED",
    ).scalar() or 0

    reject_rate = (total_rejected / total_all) if total_all else 0.0

    total_qty_approved = db.query(func.coalesce(func.sum(ProductionReportItem.qty), 0)).join(
        ProductionReport, ProductionReport.id == ProductionReportItem.report_id
    ).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "APPROVED",
    ).scalar() or 0

    totals = [
        KPIBlock(title="區間回報單數", value=int(total_all), unit="單"),
        KPIBlock(title="已核准總產量", value=int(total_qty_approved), unit="qty"),
        KPIBlock(title="退回單數", value=int(total_rejected), unit="單"),
        KPIBlock(title="退回率", value=round(reject_rate * 100, 2), unit="%", note="REJECTED / 全部回報單"),
    ]

    # --- employee rank (approved qty) ---
    # label 用 user.display_name（若找不到就顯示 ID）
    user_rows = db.query(User.id, User.display_name).all()
    user_map = {int(uid): name for uid, name in user_rows}

    emp_rows = db.query(
        ProductionReport.reported_by_user_id.label("emp_id"),
        func.coalesce(func.sum(ProductionReportItem.qty), 0).label("total_qty")
    ).join(
        ProductionReportItem, ProductionReportItem.report_id == ProductionReport.id
    ).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "APPROVED",
    ).group_by(
        ProductionReport.reported_by_user_id
    ).order_by(
        func.sum(ProductionReportItem.qty).desc()
    ).limit(top_n).all()

    employee_rank = [
        KPIRankRow(
            key=f"employee:{int(r.emp_id)}",
            label=user_map.get(int(r.emp_id), f"員工ID {int(r.emp_id)}"),
            total_qty=int(r.total_qty or 0),
        )
        for r in emp_rows
    ]

    # --- product rank (approved qty) ---
    prod_rows = db.query(
        ProductionReportItem.product_id.label("pid"),
        func.coalesce(func.sum(ProductionReportItem.qty), 0).label("total_qty")
    ).join(
        ProductionReport, ProductionReport.id == ProductionReportItem.report_id
    ).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "APPROVED",
    ).group_by(
        ProductionReportItem.product_id
    ).order_by(
        func.sum(ProductionReportItem.qty).desc()
    ).limit(top_n).all()

    prod_all = db.query(Product.id, Product.sku, Product.name).all()
    prod_map = {int(pid): (sku, name) for pid, sku, name in prod_all}

    product_rank = []
    for r in prod_rows:
        pid = int(r.pid)
        sku, name = prod_map.get(pid, ("", f"#{pid}"))
        label = f"{(sku+' - ') if sku else ''}{name}"
        product_rank.append(KPIRankRow(
            key=f"product:{pid}",
            label=label,
            total_qty=int(r.total_qty or 0),
        ))

    # --- reject reasons top (from actions) ---
    # action="REJECT" 的 comment 做聚合（只看區間內的 report）
    reject_report_ids = [rid for (rid,) in db.query(ProductionReport.id).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "REJECTED",
    ).all()]

    reasons = []
    if reject_report_ids:
        actions = db.query(ProductionReportAction.comment).filter(
            ProductionReportAction.action == "REJECT",
            ProductionReportAction.report_id.in_(reject_report_ids),
        ).all()
        cnt = {}
        for (c,) in actions:
            k = normalize_reason(c or "")
            cnt[k] = cnt.get(k, 0) + 1
        reasons = sorted(cnt.items(), key=lambda x: x[1], reverse=True)[:top_n]

    reject_reasons = [
        KPIRankRow(key=f"reason:{i}", label=label, total_qty=int(n))
        for i, (label, n) in enumerate(reasons, start=1)
    ]

    return ProductionKPIOut(
        range_from=from_date,
        range_to=to_date,
        totals=totals,
        employee_rank=employee_rank,
        product_rank=product_rank,
        reject_reasons=reject_reasons,
    )

@router.get("/production-kpi", response_model=ProductionKPIOut)
def production_kpi(
    from_date: date,
    to_date: date,
    top_n: int = 10,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.supervisor, Role.office)),
):
    """GET /production-kpi - 取得 KPI 數據"""
    return _compute_kpi(db, from_date, to_date, top_n)

@router.get("/production-kpi/export.xlsx")
def export_kpi_xlsx(
    from_date: date,
    to_date: date,
    top_n: int = 10,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.admin, Role.supervisor, Role.office)),
):
    """GET /production-kpi/export.xlsx - 匯出 KPI Excel"""
    # 直接重用 KPI 計算邏輯
    kpi = _compute_kpi(db, from_date, to_date, top_n)

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # KPI header
    ws.append(["Production KPI Export"])
    ws.append(["Range From", str(kpi.range_from)])
    ws.append(["Range To", str(kpi.range_to)])
    ws.append(["Exported At", now])
    ws.append(["Exported By", f'{user.username} ({user.role.value})'])
    ws.append([])

    # KPI blocks
    hdr_row = ws.max_row + 1  # 记录表头行号
    ws.append(["Title", "Value", "Unit", "Note"])
    for b in kpi.totals:
        ws.append([b.title, b.value, b.unit, b.note or ""])
    
    # 强化 KPI sheet 格式：表头冻结、加粗、置中
    _freeze_and_header(ws, hdr_row)
    
    # 退回率那列：找到 title == "退回率"，设置百分比格式
    for r in range(hdr_row+1, ws.max_row + 1):
        if ws.cell(r, 1).value == "退回率":
            # value 在第 2 栏，已是百分比数字（例如 12.34），套格式
            ws.cell(r, 2).number_format = '0.00'
            break
    
    _autosize(ws)

    # EmployeeRank
    ws_e = wb.create_sheet("EmployeeRank")
    ws_e.append(["Employee", "TotalQty"])
    for r in kpi.employee_rank:
        ws_e.append([r.label, r.total_qty])
    _freeze_and_header(ws_e, 1)
    _autosize(ws_e)

    # ProductRank
    ws_p = wb.create_sheet("ProductRank")
    ws_p.append(["Product", "TotalQty"])
    for r in kpi.product_rank:
        ws_p.append([r.label, r.total_qty])
    _freeze_and_header(ws_p, 1)
    _autosize(ws_p)

    # RejectReasons
    ws_r = wb.create_sheet("RejectReasons")
    ws_r.append(["Reason", "Count"])
    for r in kpi.reject_reasons:
        ws_r.append([r.label, r.total_qty])
    _freeze_and_header(ws_r, 1)
    _autosize(ws_r)

    # SummaryDaily - 已核准總產量按日
    daily_rows = db.query(
        ProductionReport.report_date.label("d"),
        func.coalesce(func.sum(ProductionReportItem.qty), 0).label("qty")
    ).join(
        ProductionReportItem, ProductionReportItem.report_id == ProductionReport.id
    ).filter(
        ProductionReport.report_date >= from_date,
        ProductionReport.report_date <= to_date,
        ProductionReport.status == "APPROVED",
    ).group_by(
        ProductionReport.report_date
    ).order_by(
        ProductionReport.report_date.asc()
    ).all()

    ws_d = wb.create_sheet("SummaryDaily")
    ws_d.append(["Date", "ApprovedTotalQty"])
    for d, qty in daily_rows:
        ws_d.append([str(d), int(qty or 0)])
    _freeze_and_header(ws_d, 1)
    _autosize(ws_d)

    # === Audit log ===
    try:
        db.add(ProductionReportAction(
            report_id=0,
            action="EXPORT_KPI",
            actor_user_id=user.id,
            actor_role=user.role.value if hasattr(user.role, 'value') else str(user.role),
            comment=f"from={from_date} to={to_date} top_n={top_n}",
        ))
        db.commit()
    except Exception:
        db.rollback()

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"production_kpi_{from_date}_{to_date}.xlsx"
    return Response(
        content=out.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


from fastapi import APIRouter, Depends, Query, Response, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from datetime import date
from typing import Optional
import io

from .db import get_db
from .models import SalesOrder, SalesOrderItem, Product, Customer
from .deps import get_current_user
from .schemas import (
    SRProductCustomersOut, SRProductCustomerRow,
    SRProductRankOut, SRProductRankRow,
    SRCustomerHistoryOut, SRCustomerStats, SOView,
)
from .models import Customer, SalesOrder, SalesOrderItem
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/sales-reports", tags=["sales-reports"])

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _apply_date_filter(q, date_from: Optional[date], date_to: Optional[date]):
    # 以 doc_date 為主（你們 ERP 用這個最直覺）
    if date_from:
        q = q.filter(SalesOrder.doc_date >= date_from)
    if date_to:
        q = q.filter(SalesOrder.doc_date <= date_to)
    return q

def _so_to_view(db: Session, so: SalesOrder) -> SOView:
    """将 SalesOrder 转换为 SOView（用于客户历史查询）"""
    prod_map = {p.id: p for p in db.query(Product).all()}
    
    # 獲取客戶資訊（地址和電話）
    customer = db.query(Customer).filter(Customer.name == so.customer_name).first()
    customer_address = customer.address if customer else None
    customer_phone = customer.phone if customer else None
    
    items = []
    for it in so.items:
        p = prod_map.get(it.product_id)
        items.append({
            "id": it.id,
            "product_id": it.product_id,
            "product_sku": getattr(p, "sku", None) if p else None,
            "product_name": getattr(p, "name", f"#{it.product_id}") if p else f"#{it.product_id}",
            "product_spec": getattr(p, "spec", None) if p else None,
            "qty": float(it.qty),
            "unit": it.unit,
            "unit_price": float(getattr(it, "unit_price", 0) or 0),
            "price_unit": getattr(it, "price_unit", "件") or "件",
            "pieces_per_case": getattr(p, "pieces_per_case", None) if p else None,
            "mark": getattr(it, "mark", None) or "",
            "note": it.note or "",
        })
    return SOView(
        id=so.id,
        so_no=so.so_no,
        customer_name=so.customer_name,
        customer_address=customer_address,
        customer_phone=customer_phone,
        doc_date=so.doc_date,
        note=so.note,
        status=so.status,
        created_at=so.created_at,
        picked_at=so.picked_at,
        picked_by_id=so.picked_by_id,
        shipped_at=so.shipped_at,
        shipped_by_id=so.shipped_by_id,
        ship_note=so.ship_note,
        logistics_no=so.logistics_no,
        items=items,
    )

@router.get("/product-customers", response_model=SRProductCustomersOut)
def product_customers(
    product_id: int = Query(..., ge=1),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    # 確認商品存在
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="product not found")

    # per-customer aggregates
    rows_q = (
        db.query(
            SalesOrder.customer_name.label("customer_name"),
            func.count(func.distinct(SalesOrder.id)).label("order_count"),
            func.coalesce(func.sum(SalesOrderItem.qty), 0).label("total_qty"),
            func.coalesce(func.sum(SalesOrderItem.qty * SalesOrderItem.unit_price), 0).label("total_amount"),
            func.max(SalesOrder.doc_date).label("last_order_date"),
        )
        .select_from(SalesOrderItem)
        .join(SalesOrder, SalesOrder.id == SalesOrderItem.sales_order_id)
        .filter(SalesOrderItem.product_id == product_id)
    )
    rows_q = _apply_date_filter(rows_q, date_from, date_to)
    rows_q = rows_q.group_by(SalesOrder.customer_name).order_by(desc("total_amount"))

    agg = rows_q.all()

    # 找每個客戶「最近一次」成交的 unit_price / price_unit（用 doc_date + id 作 tie-break）
    result_rows: list[SRProductCustomerRow] = []
    for r in agg:
        cname = (r.customer_name or "").strip()

        last_item = (
            db.query(SalesOrderItem, SalesOrder)
            .join(SalesOrder, SalesOrder.id == SalesOrderItem.sales_order_id)
            .filter(SalesOrderItem.product_id == product_id)
            .filter(SalesOrder.customer_name == cname)
        )
        last_item = _apply_date_filter(last_item, date_from, date_to)
        last_item = last_item.order_by(desc(SalesOrder.doc_date), desc(SalesOrder.id), desc(SalesOrderItem.id)).first()

        last_unit_price = 0.0
        last_price_unit = "件"
        if last_item:
            it, so = last_item
            last_unit_price = float(it.unit_price or 0)
            last_price_unit = (getattr(it, "price_unit", None) or it.unit or "件").strip() or "件"

        result_rows.append(
            SRProductCustomerRow(
                customer_name=cname,
                order_count=int(r.order_count or 0),
                total_qty=float(r.total_qty or 0),
                total_amount=float(r.total_amount or 0),
                last_unit_price=last_unit_price,
                last_price_unit=last_price_unit,
                last_order_date=r.last_order_date,
            )
        )

    return {
        "product_id": product_id,
        "date_from": date_from,
        "date_to": date_to,
        "rows": result_rows,
    }

@router.get("/product-customers/export.xlsx")
def product_customers_export_xlsx(
    product_id: int = Query(..., ge=1),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    data = product_customers(product_id, date_from, date_to, db, user)

    wb = Workbook()
    ws = wb.active
    ws.title = "ProductCustomers"
    ws.append(["product_id", "date_from", "date_to"])
    ws.append([data["product_id"], str(data["date_from"] or ""), str(data["date_to"] or "")])
    ws.append([])
    ws.append(["customer_name", "order_count", "total_qty", "total_amount", "last_unit_price", "last_price_unit", "last_order_date"])

    for r in data["rows"]:
        ws.append([
            r.customer_name,
            r.order_count,
            r.total_qty,
            r.total_amount,
            r.last_unit_price,
            r.last_price_unit,
            str(r.last_order_date or ""),
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"product_customers_{product_id}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/products-rank", response_model=SRProductRankOut)
def products_rank(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    top_n: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    q = (
        db.query(
            Product.id.label("product_id"),
            Product.sku.label("sku"),
            Product.name.label("name"),
            func.count(func.distinct(SalesOrder.id)).label("order_count"),
            func.count(func.distinct(SalesOrder.customer_name)).label("customer_count"),
            func.coalesce(func.sum(SalesOrderItem.qty), 0).label("total_qty"),
            func.coalesce(func.sum(SalesOrderItem.qty * SalesOrderItem.unit_price), 0).label("total_amount"),
        )
        .join(SalesOrderItem, SalesOrderItem.product_id == Product.id)
        .join(SalesOrder, SalesOrder.id == SalesOrderItem.sales_order_id)
    )
    q = _apply_date_filter(q, date_from, date_to)
    q = q.group_by(Product.id, Product.sku, Product.name).order_by(desc("total_amount")).limit(top_n)

    rows = []
    for r in q.all():
        rows.append(
            SRProductRankRow(
                product_id=int(r.product_id),
                sku=r.sku,
                name=r.name,
                order_count=int(r.order_count or 0),
                customer_count=int(r.customer_count or 0),
                total_qty=float(r.total_qty or 0),
                total_amount=float(r.total_amount or 0),
            )
        )

    return {"date_from": date_from, "date_to": date_to, "top_n": top_n, "rows": rows}

@router.get("/products-rank/export.xlsx")
def products_rank_export_xlsx(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    top_n: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    data = products_rank(date_from, date_to, top_n, db, user)

    wb = Workbook()
    ws = wb.active
    ws.title = "ProductsRank"
    ws.append(["date_from", "date_to", "top_n"])
    ws.append([str(data["date_from"] or ""), str(data["date_to"] or ""), data["top_n"]])
    ws.append([])
    ws.append(["product_id", "sku", "name", "order_count", "customer_count", "total_qty", "total_amount"])

    for r in data["rows"]:
        ws.append([
            r.product_id,
            r.sku or "",
            r.name,
            r.order_count,
            r.customer_count,
            r.total_qty,
            r.total_amount,
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "products_rank.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/customer-history", response_model=SRCustomerHistoryOut)
def customer_history(
    customer_name: str = Query(..., description="客户名称"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """获取客户的销货单历史和统计数据"""
    # 确认客户存在
    customer = db.query(Customer).filter(Customer.name == customer_name).first()
    if not customer:
        raise HTTPException(status_code=404, detail="customer not found")
    
    # 基础查询：该客户的所有销货单
    base_q = db.query(SalesOrder).filter(SalesOrder.customer_name == customer_name)
    
    # 应用日期过滤
    if date_from:
        base_q = base_q.filter(SalesOrder.doc_date >= date_from)
    if date_to:
        base_q = base_q.filter(SalesOrder.doc_date <= date_to)
    
    # 获取所有订单（用于统计）
    all_orders = base_q.all()
    
    # 分页查询
    total = len(all_orders)
    orders = base_q.order_by(desc(SalesOrder.doc_date), desc(SalesOrder.id))\
                   .offset((page - 1) * page_size)\
                   .limit(page_size)\
                   .all()
    
    # 计算统计数据
    total_orders = len(all_orders)
    total_amount = 0.0
    total_qty = 0.0
    orders_by_status = {"DRAFT": 0, "PICKED": 0, "SHIPPED": 0}
    first_order_date = None
    last_order_date = None
    
    # 本月和上月的日期范围
    today = date.today()
    this_month_start = date(today.year, today.month, 1)
    if today.month == 1:
        last_month_start = date(today.year - 1, 12, 1)
        last_month_end = date(today.year - 1, 12, 31)
    else:
        last_month_start = date(today.year, today.month - 1, 1)
        last_month_end = date(today.year, today.month, 1) - timedelta(days=1)
    
    orders_this_month = 0
    orders_last_month = 0
    amount_this_month = 0.0
    amount_last_month = 0.0
    
    for so in all_orders:
        # 计算订单金额和数量
        order_amount = 0.0
        order_qty = 0.0
        for item in so.items:
            item_amount = float(item.qty or 0) * float(item.unit_price or 0)
            order_amount += item_amount
            order_qty += float(item.qty or 0)
        
        total_amount += order_amount
        total_qty += order_qty
        
        # 按状态统计
        status = so.status or "DRAFT"
        if status in orders_by_status:
            orders_by_status[status] += 1
        
        # 首次和最近订单日期
        if so.doc_date:
            if first_order_date is None or so.doc_date < first_order_date:
                first_order_date = so.doc_date
            if last_order_date is None or so.doc_date > last_order_date:
                last_order_date = so.doc_date
            
            # 本月和上月统计
            if so.doc_date >= this_month_start:
                orders_this_month += 1
                amount_this_month += order_amount
            elif last_month_start <= so.doc_date <= last_month_end:
                orders_last_month += 1
                amount_last_month += order_amount
    
    avg_order_amount = total_amount / total_orders if total_orders > 0 else 0.0
    
    stats = SRCustomerStats(
        total_orders=total_orders,
        total_amount=total_amount,
        total_qty=total_qty,
        avg_order_amount=avg_order_amount,
        first_order_date=first_order_date,
        last_order_date=last_order_date,
        orders_by_status=orders_by_status,
        orders_this_month=orders_this_month,
        orders_last_month=orders_last_month,
        amount_this_month=amount_this_month,
        amount_last_month=amount_last_month,
    )
    
    return SRCustomerHistoryOut(
        customer_name=customer_name,
        date_from=date_from,
        date_to=date_to,
        stats=stats,
        orders=[_so_to_view(db, so) for so in orders],
    )

